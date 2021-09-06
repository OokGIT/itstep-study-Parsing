from openpyxl import Workbook, open

xls_file = 'example.xlsx'
xls_sheet = 'from_my_dict'


class ListIteration:

    def __init__(self, a):
        print("Наш список для итераций: \n", a)
        # print('init')
        self.a = a
        self.my_dict = {}
        self.wb = Workbook()
        self.ws = self.wb.create_sheet(xls_sheet)
        self.wb.active = 1
        # print(self.wb.active)

    def unpack_list(self):
        print("Разворачиваем список:")
        for i in self.a:
            print(i, end=' ')
        print()

    def dictate(self):
        # self.my_dict = {}
        print("Собрали из списка такой dict:")
        for i in self.a:
            self.my_dict[i] = i
        print(self.my_dict)

    def write_to_xls(self):
        if self.my_dict:
            # print('дикт для распаковки в xls:', self.my_dict)
            for key in self.my_dict.keys():
                self.ws.cell(row=key, column=1).value = self.my_dict[key]
            self.wb.save(xls_file)
            self.wb.close()

    def from_xls(self):
        file = open(xls_file, read_only=True)
        sheet = file.active
        for row in range(1, sheet.max_row):
            value = sheet[row][0].value
            print("В строке", row, "Соджержится значение:", value)
        file.close()


obj_a = ListIteration(a=[1, 2, 3, 12, 8, 4])
obj_a.unpack_list()
obj_a.dictate()
obj_a.write_to_xls()
obj_a.from_xls()
# obj_a.func()
# obj_a._func()
